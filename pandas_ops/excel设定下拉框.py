#!/usr/bin/env python3
# -*- encoding: utf-8 -*-


import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

# 创建一个新的Excel工作簿
workbook = openpyxl.Workbook()
worksheet = workbook.active

# 假设我们要在第2列（B列）添加下拉框
column_index = 2

# 定义下拉框的选项
options = ['选项1', '选项2', '选项3']

# 创建一个数据验证对象
dv = DataValidation(type="list", formula1=f'"{",".join(options)}"', showDropDown=True)

# 将数据验证应用到指定的列
for row in range(1, 10):  # 假设我们要应用到第1行到第10行
    cell = worksheet.cell(row=row, column=column_index)
    dv.add(cell)

# 设置默认值
default_value = '选项1'
worksheet.cell(row=1, column=column_index).value = default_value

# 将数据验证添加到工作表
worksheet.add_data_validation(dv)

# 保存工作簿
workbook.save('example.xlsx')
