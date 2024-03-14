#!/usr/bin/env python3
# -*- coding: utf8 -*-
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def set_style(writer, sheet):
    workbook = writer.book
    worksheet = workbook[sheet]
    # 列宽自适应
    for column in worksheet.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 2
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    # 设置列名的底色
    header_fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type="solid")
    for cell in worksheet[1]:
        cell.fill = header_fill